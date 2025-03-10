from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import pandas as pd
import matplotlib.pyplot as plt
import io
import base64
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
import os
import seaborn as sns
import logging
from logging.handlers import RotatingFileHandler
from dashboard import dash_app
from dash import Dash
from flask import Flask, render_template
from flask import Flask
from app.routes.diagnostic import diagnostic_bp
from app.routes.data_routes import data_bp

app = Flask(__name__)
app.register_blueprint(diagnostic_bp)
app.register_blueprint(data_bp, url_prefix='/')

if __name__ == '__main__':
    app.run(debug=True)

# Configuração do Google Sheets
SCOPE = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
CREDENTIALS_PATH = os.path.join(os.path.dirname(__file__), "..", "google_sheets", "credentials.json")
CREDS = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_PATH, SCOPE)
CLIENT = gspread.authorize(CREDS)
SHEET = CLIENT.open("RespostasForm")

# Mapeamento dos formulários para suas abas correspondentes
FORM_SHEETS = {
    "Recanto das Emas": SHEET.worksheet("Recanto das Emas"),
    "Gama": SHEET.worksheet("Gama"),
    "Santa Maria": SHEET.worksheet("Santa Maria"),
    "Guara": SHEET.worksheet("Guara"),
    "Planaltina": SHEET.worksheet("Planaltina"),
    "Samambaia": SHEET.worksheet("Samambaia"),
}

def setup_logger():
    logger = logging.getLogger('app')
    logger.setLevel(logging.INFO)
    handler = RotatingFileHandler('app.log', maxBytes=10000, backupCount=3)
    formatter = logging.Formatter(
        '[%(asctime)s] %(levelname)s in %(module)s: %(message)s'
    )
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    return logger

logger = setup_logger()

@app.route("/")
def index():
    return redirect('/dashboard/')

@app.route("/visualizar")
def gerar_grafico():
    """
    Gera visualizações gráficas dos dados coletados.
    
    Returns:
        str: Template HTML renderizado com os gráficos gerados
        
    Raises:
        Exception: Se houver erro na geração dos gráficos
    """
    try:
        dados_consolidados = []
        graficos = {}

        # Coleta dados de todas as planilhas
        for form, planilha in FORM_SHEETS.items():
            registros = planilha.get_all_records()
            for registro in registros:
                dados_registro = {
                    'deficiencia': None,
                    'atendimento': 'Atendido',  # Valor padrão para contagem
                    'interesse_politica': None,
                    'pontos_melhoria': None,
                    'cidade': form  # Adiciona a cidade (que é o nome do form)
                }
                
                for coluna in registro.keys():
                    # Procura por colunas relacionadas à deficiência
                    if "4 - Qual" in coluna and "deficiência" in coluna.lower():
                        dados_registro['deficiencia'] = registro[coluna]
                    # Procura por colunas relacionadas ao interesse em política
                    elif any(termo in coluna.lower() for termo in ["política", "politica"]):
                        dados_registro['interesse_politica'] = registro[coluna]
                    # Procura por colunas relacionadas a pontos de melhoria
                    elif any(termo in coluna.lower() for termo in ["melhorar", "melhoria"]):
                        dados_registro['pontos_melhoria'] = registro[coluna]
                
                dados_consolidados.append(dados_registro)

        if not dados_consolidados:
            return "<h1>Nenhum dado disponível nas planilhas</h1>"

        df = pd.DataFrame(dados_consolidados)
        print("Colunas encontradas:", df.columns.tolist())  # Debug

        # Configurar um estilo mais profissional para os gráficos
        plt.style.use('seaborn')
        sns.set_palette("husl")  # Paleta de cores mais profissional

        # Adicionar formatação mais detalhada aos gráficos
        plt.figure(figsize=(12, 6))
        df_def = df['deficiencia'].value_counts()
        sns.barplot(x=df_def.index, y=df_def.values)
        plt.title('Quantidade de Pessoas por Deficiência', pad=20, fontsize=14)
        plt.xlabel('Tipo de Deficiência', fontsize=12)
        plt.ylabel('Quantidade', fontsize=12)
        plt.grid(True, alpha=0.3)
        img = io.BytesIO()
        plt.savefig(img, format='png', bbox_inches='tight')
        img.seek(0)
        graficos["Tipos de Deficiência"] = base64.b64encode(img.getvalue()).decode()
        plt.close()

        # Gráfico 2: Atendimentos
        plt.figure(figsize=(8, 8))
        total_atendidos = len(df)
        plt.pie([total_atendidos], labels=['Atendidos'], autopct='%1.1f%%', colors=['#2ecc71'])
        plt.title('Pessoas Atendidas')
        img = io.BytesIO()
        plt.savefig(img, format='png', bbox_inches='tight')
        img.seek(0)
        graficos["Atendimentos"] = base64.b64encode(img.getvalue()).decode()
        plt.close()

        # Gráfico de Interesse em Política
        if 'interesse_politica' in df.columns and df['interesse_politica'].notna().any():
            plt.figure(figsize=(15, 12))
            
            # Subplot 1: Gráfico total de Sim/Não
            plt.subplot(3, 1, 1)
            total_interesse = df['interesse_politica'].value_counts()
            ax = sns.barplot(x=total_interesse.index, y=total_interesse.values)
            plt.title('Interesse em Política - Total Geral', pad=20, fontsize=14)
            plt.xlabel('Resposta', fontsize=12)
            plt.ylabel('Quantidade', fontsize=12)
            
            # Adicionar valores nas barras
            for i, v in enumerate(total_interesse.values):
                plt.text(i, v, str(v), ha='center', va='bottom')
            
            # Subplot 2: Gráfico por cidade
            plt.subplot(3, 1, 2)
            df_cidade = pd.crosstab(df['cidade'], df['interesse_politica'])
            ax = df_cidade.plot(kind='bar', ax=plt.gca())
            plt.title('Interesse em Política por Cidade', pad=20, fontsize=14)
            plt.xlabel('Cidade', fontsize=12)
            plt.ylabel('Quantidade', fontsize=12)
            plt.legend(title='Resposta')
            plt.xticks(rotation=45)
            
            # Adicionar valores em cada barra
            for container in ax.containers:
                ax.bar_label(container)
            
            # Subplot 3: Tabela com valores
            plt.subplot(3, 1, 3)
            plt.axis('off')
            
            # Criar tabela com os dados
            cell_text = []
            rows = []
            for cidade in df_cidade.index:
                row = [
                    cidade,
                    df_cidade.loc[cidade, 'Sim'] if 'Sim' in df_cidade.columns else 0,
                    df_cidade.loc[cidade, 'Não'] if 'Não' in df_cidade.columns else 0,
                    df_cidade.loc[cidade].sum()
                ]
                cell_text.append(row)
                rows.append(cidade)
            
            # Adicionar linha com totais
            totals = ['Total',
                     df_cidade['Sim'].sum() if 'Sim' in df_cidade.columns else 0,
                     df_cidade['Não'].sum() if 'Não' in df_cidade.columns else 0,
                     df_cidade.values.sum()]
            cell_text.append(totals)
            
            table = plt.table(cellText=cell_text,
                            colLabels=['Cidade', 'Sim', 'Não', 'Total'],
                            loc='center',
                            cellLoc='center')
            table.auto_set_font_size(False)
            table.set_fontsize(9)
            table.scale(1.2, 1.5)
            
            # Ajustar layout
            plt.tight_layout()
            
            # Salvar gráfico
            img = io.BytesIO()
            plt.savefig(img, format='png', bbox_inches='tight', dpi=300)
            img.seek(0)
            graficos["Interesse em Política"] = base64.b64encode(img.getvalue()).decode()
            plt.close()

        # Gráfico 4: Pontos de Melhoria
        if 'pontos_melhoria' in df.columns and df['pontos_melhoria'].notna().any():
            plt.figure(figsize=(10, 6))
            df_melh = df['pontos_melhoria'].value_counts()
            sns.barplot(x=df_melh.index, y=df_melh.values)
            plt.xticks(rotation=45, ha='right')
            plt.title('Pontos que Precisam de Melhoria')
            plt.xlabel('Área')
            plt.ylabel('Quantidade')
            img = io.BytesIO()
            plt.savefig(img, format='png', bbox_inches='tight')
            img.seek(0)
            graficos["Pontos de Melhoria"] = base64.b64encode(img.getvalue()).decode()
            plt.close()

        return render_template("visualizar.html", graficos=graficos)

    except Exception as e:
        import traceback
        erro = traceback.format_exc()
        print(f"Erro ao gerar gráfico: {erro}")
        return f"<pre>Erro: {erro}</pre>"

@app.route('/diagnostico-pagina')
def diagnostico_pagina():
    dados = [
        {
            "1 - Qual seu nome?": "João",
            "2 - Qual seu Telefone?": "123456789",
            "3 - Cidade": "São Paulo",
            "4 - Qual o tipo de deficiência?": "Visual",
            "5 - Você conhece todos os atendimentos e serviços oferecidos pela SEPD (Secretaria da Pessoa com Deficiência)?": "Não",
            "6 - Qual benefício você precisa?": "Acessibilidade",
            "7 - Você ficou sabendo da Carreta da Inclusão?": "Sim",
            "8 - Se sim, por quem?": "Amigo",
            "9 - Se interessa por política?": "Sim",
            "Carimbo de data/hora": "2025-02-28 10:00:00"
        },
        # Mais registros...
    ]
    return render_template('diagnostico.html', dados=dados)

if __name__ == '__main__':
    app.run(debug=True)

@app.route("/dados")
def visualizar_dados():
    try:
        todos_dados = []
        colunas_ordenadas = [
            "Origem",
            "Qual seu nome",
            "Qual seu telefone",
            "Cidade",
            "Qual o tipo de deficiência",
            "Você conhece todos os atendimentos e serviços oferecidos pela SEPD (Secretaria da Pessoa com Deficiência)?",
            "Qual benefício você precisa",
            "Você ficou sabendo da Carreta da Inclusão",
            "Se sim, por quem",
            "Se interessa por Política"
        ]

        for form_name, worksheet in FORM_SHEETS.items():
            registros = worksheet.get_all_records()
            for registro in registros:
                registro['Origem'] = form_name
                dado_formatado = {'Origem': form_name}
                for coluna in colunas_ordenadas[1:]:
                    valor = None
                    for chave in registro.keys():
                        if coluna.lower() in chave.lower():
                            valor = registro[chave]
                            break
                    dado_formatado[coluna] = valor
                todos_dados.append(dado_formatado)

        return render_template("dados.html", todos_dados=todos_dados)
    except Exception as e:
        import traceback
        erro = traceback.format_exc()
        return f"<pre>Erro ao carregar dados: {erro}</pre>"

@app.route("/exportar_excel")
def exportar_excel():
    try:
        todos_dados = []
        colunas_ordenadas = [
            "Origem",
            "Qual seu nome",
            "Qual seu telefone",
            "Cidade",
            "Qual o tipo de deficiência",
            "Você conhece todos os atendimentos e serviços oferecidos pela SEPD (Secretaria da Pessoa com Deficiência)?",
            "Qual benefício você precisa",
            "Você ficou sabendo da Carreta da Inclusão",
            "Se sim, por quem",
            "Se interessa por Política"
        ]

        for form_name, worksheet in FORM_SHEETS.items():
            registros = worksheet.get_all_records()
            for registro in registros:
                dado_formatado = {'Origem': form_name}
                for coluna in colunas_ordenadas[1:]:
                    valor = None
                    for chave in registro.keys():
                        if coluna.lower() in chave.lower():
                            valor = registro[chave]
                            break
                    dado_formatado[coluna] = valor
                todos_dados.append(dado_formatado)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dados Formulários"

        header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col, header in enumerate(colunas_ordenadas, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row, dado in enumerate(todos_dados, 2):
            for col, coluna in enumerate(colunas_ordenadas, 1):
                ws.cell(row=row, column=col, value=dado.get(coluna, ''))

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"dados_formularios_{timestamp}.xlsx"
        filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)
        
        response = send_file(
            filepath,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
        @response.call_on_close
        def remove_file():
            try:
                os.remove(filepath)
            except:
                pass
        
        return response

    except Exception as e:
        import traceback
        erro = traceback.format_exc()
        return f"Erro ao exportar dados: {erro}", 500

@app.route("/analytics")
def analytics():
    try:
        # Análise de tendências
        dados_por_mes = df.groupby(pd.Grouper(freq='M')).count() # type: ignore
        
        # Análise comparativa entre regiões
        comparativo_regioes = df.groupby('Origem').agg({ # type: ignore
            'deficiencia': 'count',
            'interesse_politica': lambda x: (x == 'Sim').sum()
        })
        
        # Calcular métricas importantes
        total_atendimentos = len(df) # type: ignore
        media_por_regiao = df.groupby('Origem').size().mean() # type: ignore
        
        return render_template(
            "analytics.html",
            metricas={
                'total_atendimentos': total_atendimentos,
                'media_por_regiao': media_por_regiao,
                'dados_por_mes': dados_por_mes.to_dict(),
                'comparativo_regioes': comparativo_regioes.to_dict()
            }
        )
    except Exception as e:
        return f"Erro na análise: {str(e)}"

@app.route("/exportar_relatorio")
def exportar_relatorio():
    try:
        # Criar relatório PDF com ReportLab ou WeasyPrint
        # Incluir gráficos, tabelas e análises
        # Adicionar marca d'água e cabeçalho oficial
        pass
    except Exception as e:
        return f"Erro ao gerar relatório: {str(e)}"

@app.route("/api/dados", methods=['GET'])
def api_dados():
    try:
        dados = {
            'total_atendimentos': len(df), # type: ignore
            'por_regiao': df.groupby('Origem').size().to_dict(), # type: ignore
            'por_deficiencia': df['deficiencia'].value_counts().to_dict() # type: ignore
        }
        return jsonify(dados)
    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/dashboard')
def dashboard():
    return dash_app.index()

if __name__ == "__main__":
    app.run(debug=True)

    #http://127.0.0.1:8080
    #python run.py
    #.\venv\Scripts\activate